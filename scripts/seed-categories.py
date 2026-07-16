"""
Seed script: gera `src/data/categories.json` a partir de:
  1. data/Categorias-ajustada-pos-sugestoes-sites-09-07-26-PROPOSTA2-v8.xlsx
     (sheet 'Categorias e Subcategorias') — taxonomia oficial: categoria, subcategoria,
     título da carta e ÓRGÃO (sigla). Fonte de verdade do mapeamento. A coluna Órgão é o
     que permite diferenciar serviços comuns a vários órgãos (ouvidoria, LAI, peticionamento,
     vista e cópia): cada órgão vira uma carta com a sua própria página.
  2. data/cartas-ativas-com-url.xlsx (sheet 'cartas_ativas') — lookup da URL por
     (título, sigla do órgão). É a lista viva de páginas do portal.
  3. data/Todas as cartas.xlsx (sheets 'Proposta 2' + 'servicos_ativos (1)') — apenas para
     mapear a sigla do órgão no nome completo exibido no card.
  4. data/subcategoria-rules.json — metadados por categoria (id slug + ícone Material).

Match por (título, sigla) normalizados (sem acentos, lowercase). Quando a sigla diverge,
tenta fuzzy no título dentro do mesmo órgão (difflib, cutoff 0.9).

Categorias e subcategorias saem em ordem alfabética.

Run:  npm run seed     (ou:  python scripts/seed-categories.py)
"""
from __future__ import annotations

import difflib
import json
import re
import sys
import unicodedata
from collections import defaultdict
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parent.parent
TAXONOMY_PATH = (
    ROOT / "data" / "Categorias-ajustada-pos-sugestoes-sites-09-07-26-PROPOSTA2-v8.xlsx"
)
ACTIVE_PATH = ROOT / "data" / "cartas-ativas-com-url.xlsx"
LEGACY_PATH = ROOT / "data" / "Todas as cartas.xlsx"
META_PATH = ROOT / "data" / "subcategoria-rules.json"
OUTPUT_PATH = ROOT / "src" / "data" / "categories.json"

TAXONOMY_SHEET = "Categorias e Subcategorias"
ACTIVE_SHEET = "cartas_ativas"
# Ordem importa para o nome do órgão: a primeira aba que tiver a sigla vence.
LEGACY_SHEETS = ("Proposta 2", "servicos_ativos (1)")

# Fuzzy só entra quando a sigla diverge; cutoff alto para não trocar serviços parecidos.
FUZZY_CUTOFF = 0.90

# Renomeações de subcategoria aplicadas por cima da planilha (chave = nome na planilha).
SUBCATEGORY_RENAMES = {
    "Contencioso": "Processos Administrativos Tributários",
}


def normalize(text: str | None) -> str:
    if not text:
        return ""
    text = unicodedata.normalize("NFKD", str(text)).encode("ascii", "ignore").decode()
    return re.sub(r"\s+", " ", text).strip().lower()


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = re.sub(r"[^a-zA-Z0-9]+", "-", text).strip("-").lower()
    return text or "item"


def clean(text: object) -> str:
    """Collapse runs of whitespace — algumas células da planilha têm espaço duplo."""
    return re.sub(r"\s+", " ", str(text)).strip()


def load_taxonomy() -> list[tuple[str, str, str, str]]:
    """Read the taxonomy. Returns [(categoria, subcategoria, titulo, sigla), ...] com
    cascade-fill nas células vazias de categoria/subcategoria (estilo Excel mesclado)."""
    wb = openpyxl.load_workbook(TAXONOMY_PATH, data_only=True, read_only=True)
    ws = wb[TAXONOMY_SHEET]
    rows: list[tuple[str, str, str, str]] = []
    current_cat, current_sub = None, None
    for row in ws.iter_rows(min_row=4, values_only=True):
        cat, sub, titulo, orgao = row[0], row[1], row[2], row[3]
        if cat:
            current_cat = clean(cat)
        if sub:
            current_sub = clean(sub)
            current_sub = SUBCATEGORY_RENAMES.get(current_sub, current_sub)
        if titulo and current_cat and current_sub:
            rows.append((current_cat, current_sub, clean(titulo), clean(orgao) if orgao else ""))
    return rows


def load_url_lookup() -> tuple[dict[tuple[str, str], str], dict[str, list[str]]]:
    """cartas-ativas-com-url.xlsx → índice da URL por (título, sigla).

    Retorna:
      by_ts:            {(norm_título, norm_sigla): url}
      titles_by_sigla:  {norm_sigla: [norm_título, ...]}  (base do fuzzy por órgão)
    """
    ws = openpyxl.load_workbook(ACTIVE_PATH, data_only=True, read_only=True)[ACTIVE_SHEET]
    by_ts: dict[tuple[str, str], str] = {}
    titles_by_sigla: dict[str, list[str]] = defaultdict(list)
    for row in ws.iter_rows(min_row=2, values_only=True):
        sigla = clean(row[0]) if row[0] else ""
        titulo = clean(row[1]) if row[1] else ""
        url = clean(row[2]) if row[2] else ""
        if not titulo or not url:
            continue
        key = (normalize(titulo), normalize(sigla))
        if key not in by_ts:
            by_ts[key] = url
            titles_by_sigla[normalize(sigla)].append(normalize(titulo))
    return by_ts, titles_by_sigla


def load_orgao_names() -> dict[str, str]:
    """Todas as cartas.xlsx → {norm_sigla: nome completo do órgão} para exibição."""
    wb = openpyxl.load_workbook(LEGACY_PATH, data_only=True, read_only=True)
    names: dict[str, str] = {}
    for sheet_name in LEGACY_SHEETS:
        if sheet_name not in wb.sheetnames:
            continue
        for row in wb[sheet_name].iter_rows(min_row=2, values_only=True):
            sigla, _titulo, _cat, _url, orgao = row[:5]
            key = normalize(sigla)
            if key and orgao and key not in names:
                names[key] = clean(orgao)
    return names


def main() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")

    print(f"[seed] Lendo taxonomia: {TAXONOMY_PATH.relative_to(ROOT)} (sheet: {TAXONOMY_SHEET})")
    records = load_taxonomy()
    print(f"[seed]   {len(records)} cartas com (categoria, subcategoria, título, órgão)")

    print(f"[seed] Lendo URLs por (título, órgão): {ACTIVE_PATH.relative_to(ROOT)}")
    by_ts, titles_by_sigla = load_url_lookup()
    print(f"[seed]   {len(by_ts)} páginas ativas")

    print(f"[seed] Lendo nomes de órgão: {LEGACY_PATH.relative_to(ROOT)}")
    orgao_names = load_orgao_names()

    print(f"[seed] Lendo metadados de categoria: {META_PATH.relative_to(ROOT)}")
    meta = json.loads(META_PATH.read_text(encoding="utf-8"))

    grouped: dict[str, list[tuple[str, dict]]] = defaultdict(list)
    exact_hits = 0
    fuzzy_hits = 0
    no_match = 0
    no_url: list[tuple[str, str, str, str]] = []
    seen_ids: dict[str, int] = {}

    unknown_cats = sorted({c for c, _, _, _ in records} - set(meta))
    if unknown_cats:
        print(f"[seed] AVISO: categoria(s) fora de {META_PATH.name} "
              f"(id/ícone gerados por slug): {', '.join(unknown_cats)}")

    for cat_name, sub_name, titulo, sigla in records:
        nt, ns = normalize(titulo), normalize(sigla)
        url = by_ts.get((nt, ns), "")
        if url:
            exact_hits += 1
        else:
            # Sigla diverge: procura o título mais próximo dentro do mesmo órgão.
            near = difflib.get_close_matches(nt, titles_by_sigla.get(ns, []), n=1, cutoff=FUZZY_CUTOFF)
            if near:
                url = by_ts.get((near[0], ns), "")
            if url:
                fuzzy_hits += 1
            else:
                no_match += 1

        if not url:
            no_url.append((cat_name, sub_name, titulo, sigla))

        agency = orgao_names.get(ns) or sigla

        cat_id = meta.get(cat_name, {}).get("id") or slugify(cat_name)
        sub_id = slugify(sub_name)
        card_id = f"{cat_id}--{sub_id}--{slugify(f'{titulo}-{sigla}')[:90]}"
        # Mesma (carta, órgão) repetida na subcategoria: sufixo para manter o id único.
        seen_ids[card_id] = seen_ids.get(card_id, 0) + 1
        if seen_ids[card_id] > 1:
            card_id = f"{card_id}-{seen_ids[card_id]}"

        grouped[cat_name].append((sub_name, {
            "id": card_id,
            "title": titulo,
            "url": url,
            "agency": agency,
            "agencyCode": sigla,
        }))

    # Categorias e subcategorias em ordem alfabética (ignorando acentos).
    output = []
    for cat_name in sorted(grouped, key=normalize):
        items = grouped[cat_name]
        cat_meta = meta.get(cat_name) or {"id": slugify(cat_name), "icon": "category"}

        sub_cards: dict[str, list[dict]] = defaultdict(list)
        for sub_name, card in items:
            sub_cards[sub_name].append(card)

        subcategories_out = []
        for sub_name in sorted(sub_cards, key=normalize):
            cards = sorted(sub_cards[sub_name], key=lambda c: (normalize(c["title"]), c["agencyCode"]))
            subcategories_out.append({
                "id": slugify(sub_name),
                "name": sub_name,
                "count": len(cards),
                "cards": cards,
            })

        total = sum(s["count"] for s in subcategories_out)
        output.append({
            "id": cat_meta["id"],
            "name": cat_name,
            "icon": cat_meta["icon"],
            "count": total,
            "subcategories": subcategories_out,
        })

    # Report
    print()
    print("[seed] === Resultado ===")
    for cat in output:
        print(f"  {cat['count']:>4}  {cat['name']}")
        for sub in cat["subcategories"]:
            print(f"        {sub['count']:>4}  +- {sub['name']}")
    total_cards = sum(c["count"] for c in output)
    print()
    print(f"[seed] Total: {total_cards} cartas")
    matched = exact_hits + fuzzy_hits
    print(f"[seed] URL match: {exact_hits} exato + {fuzzy_hits} fuzzy "
          f"= {matched} ({100 * matched // max(total_cards, 1)}%); {no_match} sem match")
    if no_url:
        print()
        print(f"[seed] {len(no_url)} carta(s) SEM URL (link vazio):")
        for cat_name, sub_name, titulo, sigla in no_url:
            print(f"        - [{cat_name} / {sub_name}] {titulo}  ({sigla or 'sem órgão'})")

    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_PATH.write_text(json.dumps(output, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"[seed] Gravado: {OUTPUT_PATH.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
