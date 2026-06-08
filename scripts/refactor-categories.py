"""
Migração das categorias conforme planilha (6.2) Linguagem Simples v2.xlsx.
Gera novo src/data/categories.json com nova estrutura de 13 categorias / 69 subcategorias.
"""

import io
import json
import re
import sys
import unicodedata
from collections import OrderedDict
from pathlib import Path

import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding="utf-8")

ROOT = Path(__file__).parent.parent
EXCEL_PATH = ROOT / "data" / "(6.2) Linguagem Simples  v2.xlsx"
INPUT_JSON = ROOT / "src" / "data" / "categories.json"
OUTPUT_JSON = ROOT / "src" / "data" / "categories.json"
RULES_PATH = ROOT / "data" / "subcategoria-rules.json"

# ── Icon mapping for all 13 categories ────────────────────────────────────────
CATEGORY_ICONS = {
    "Administração Pública":                     "account_balance",
    "Agronegócio e Meio Ambiente":               "agriculture",
    "Arte, Esporte e Turismo":                   "theater_comedy",
    "Assistência Social, Habitação e Cidadania": "diversity_3",
    "Ciência, Educação e Pesquisa":              "auto_stories",
    "Comunicação e Transparência":               "campaign",
    "Empresa, Indústria e Comércio":             "apartment",
    "Finanças e Impostos":                       "currency_exchange",
    "Justiça e Segurança":                       "balance",
    "Saneamento, Água e Energia":                "water_drop",
    "Saúde e Cuidado":                           "medical_services",
    "Trabalho, Emprego e Previdência":           "work",
    "Trânsito, Transportes e Infraestrutura":    "directions_car",
}


def slugify(text: str) -> str:
    text = unicodedata.normalize("NFD", text)
    text = "".join(c for c in text if unicodedata.category(c) != "Mn")
    text = text.lower()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def normalize_title(title: str) -> str:
    # Decode Excel encoding artifacts for special characters
    title = title.replace("_x0096_", "–").replace("_x0093_", "“").replace("_x0094_", "”")
    title = unicodedata.normalize("NFD", title)
    title = "".join(c for c in title if unicodedata.category(c) != "Mn")
    # Normalize dashes and punctuation variants
    title = re.sub(r"[–—―]", "-", title)
    title = re.sub(r"[;,]", ",", title)  # treat ; and , as equivalent
    title = re.sub(r"\s+", " ", title)
    return title.lower().strip()


# ── Step 1: read Excel structure ───────────────────────────────────────────────
def read_excel() -> OrderedDict:
    wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True)
    ws = wb[wb.sheetnames[0]]

    structure: OrderedDict = OrderedDict()
    current_cat = current_sub = None

    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i < 3:
            continue
        cat, sub, carta = row[0], row[1], row[2]
        if cat:
            current_cat = str(cat).strip()
            if current_cat not in structure:
                structure[current_cat] = OrderedDict()
        if sub:
            current_sub = str(sub).strip()
            if current_cat and current_sub not in structure[current_cat]:
                structure[current_cat][current_sub] = []
        if carta and current_cat and current_sub:
            structure[current_cat][current_sub].append(str(carta).strip())

    wb.close()
    return structure


# ── Step 2: build title → card lookup from current JSON ───────────────────────
def build_lookup(data: list) -> dict:
    lookup: dict = {}
    for cat in data:
        for sub in cat["subcategories"]:
            for card in sub["cards"]:
                key = normalize_title(card["title"])
                if key in lookup:
                    lookup[key].append(card)
                else:
                    lookup[key] = [card]
    return lookup


# ── Step 3: build new categories JSON ─────────────────────────────────────────
def build_new_json(structure: OrderedDict, lookup: dict) -> tuple[list, list, list]:
    new_categories = []
    not_found: list[tuple[str, str, str]] = []   # (cat, sub, title)

    # Convert lists to deques so we can pop from front (consuming each card once)
    from collections import deque
    pool: dict[str, deque] = {k: deque(v) for k, v in lookup.items()}

    # Track IDs generated to detect slug collisions within a subcategory
    used_ids: set[str] = set()

    for cat_name, subcats in structure.items():
        cat_id = slugify(cat_name)
        icon = CATEGORY_ICONS.get(cat_name, "help_outline")

        new_subcats = []
        for sub_name, titles in subcats.items():
            sub_id = slugify(sub_name)
            new_cards = []

            for title in titles:
                norm = normalize_title(title)
                cards_queue = pool.get(norm)
                if not cards_queue:
                    not_found.append((cat_name, sub_name, title))
                    continue

                card = cards_queue.popleft()  # consume one card from pool
                base_id = f"{cat_id}--{sub_id}--{slugify(card['title'])}"
                # Disambiguate slug collisions by appending a counter
                final_id = base_id
                suffix = 2
                while final_id in used_ids:
                    final_id = f"{base_id}-{suffix}"
                    suffix += 1
                used_ids.add(final_id)

                new_card = {"id": final_id, "title": card["title"], "url": card["url"],
                            "agency": card["agency"], "agencyCode": card["agencyCode"]}
                if card.get("modalities"):
                    new_card["modalities"] = card["modalities"]
                if card.get("audiences"):
                    new_card["audiences"] = card["audiences"]
                new_cards.append(new_card)

            new_subcats.append({
                "id": sub_id,
                "name": sub_name,
                "count": len(new_cards),
                "cards": new_cards,
            })

        total = sum(s["count"] for s in new_subcats)
        new_categories.append({
            "id": cat_id,
            "name": cat_name,
            "icon": icon,
            "count": total,
            "subcategories": new_subcats,
        })

    return new_categories, not_found, []


# ── Step 4: update subcategoria-rules.json ────────────────────────────────────
def update_rules(structure: OrderedDict) -> None:
    new_rules = {}
    for cat_name in structure:
        cat_id = slugify(cat_name)
        icon = CATEGORY_ICONS.get(cat_name, "help_outline")
        new_rules[cat_name] = {"id": cat_id, "icon": icon}

    with open(RULES_PATH, "w", encoding="utf-8") as f:
        json.dump(new_rules, f, ensure_ascii=False, indent=2)
    print(f"✔  subcategoria-rules.json atualizado ({len(new_rules)} categorias)")


# ── Main ───────────────────────────────────────────────────────────────────────
def main() -> None:
    print("Lendo planilha…")
    structure = read_excel()
    total_excel = sum(len(cards) for subs in structure.values() for cards in subs.values())
    print(f"  {len(structure)} categorias | {sum(len(s) for s in structure.values())} subcategorias | {total_excel} títulos")

    print("Lendo categories.json atual…")
    with open(INPUT_JSON, encoding="utf-8") as f:
        current_data = json.load(f)
    total_json = sum(s["count"] for cat in current_data for s in cat["subcategories"])
    print(f"  {len(current_data)} categorias | {total_json} cartas")

    print("Construindo lookup por título…")
    lookup = build_lookup(current_data)
    print(f"  {len(lookup)} títulos únicos no JSON atual")

    print("Gerando nova estrutura…")
    new_categories, not_found, duplicates = build_new_json(structure, lookup)

    new_total = sum(c["count"] for c in new_categories)
    sep = "-" * 60
    print(f"\n{sep}")
    print(f"  Categorias geradas : {len(new_categories)}")
    print(f"  Total de cartas    : {new_total}")
    print(f"  Nao encontradas    : {len(not_found)}")
    print(sep)

    if not_found:
        print("\n⚠  CARTAS NÃO ENCONTRADAS (titulo da planilha sem correspondencia no JSON):")
        for cat, sub, title in not_found:
            print(f"   [{cat} > {sub}] {title}")

    print(f"\nSalvando {OUTPUT_JSON}…")
    with open(OUTPUT_JSON, "w", encoding="utf-8") as f:
        json.dump(new_categories, f, ensure_ascii=False, indent=2)
    print("✔  categories.json salvo")

    update_rules(structure)

    print("\n✅ Migração concluída.")


if __name__ == "__main__":
    main()
